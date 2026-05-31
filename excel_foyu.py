import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os

BASE = 'D:\\00 ŞİRKET DOSYALARI\\GENEL ŞİRKET DOSYALARI\\WEB SİTESİ FOTOLARI'
KATEGORILER = {
    'konut':       'Konut Projeleri',
    'ticari':      'Ticari Projeler',
    'endustriyel': 'Endustriyel Projeler'
}

def thin():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def thick():
    s = Side(style='medium', color='F4A261')
    return Border(left=s, right=s, top=s, bottom=s)

for kat, baslik in KATEGORILER.items():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Proje Bilgileri'

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 22
    ws.column_dimensions['H'].width = 38

    # Baslik
    ws.merge_cells('A1:H1')
    c = ws['A1']
    c.value = 'MEFSTEEL HAFIF CELIK — ' + baslik.upper() + ' BILGI FOYU'
    c.font = Font(bold=True, size=14, color='FFFFFF', name='Calibri')
    c.fill = PatternFill('solid', fgColor='0D1B2A')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    ws.merge_cells('A2:H2')
    c = ws['A2']
    c.value = 'Her proje icin 1 satir doldurun. Fotograflari klasore ekleyin, GUNCELLE.bat calistirin.'
    c.font = Font(italic=True, size=9, color='555555')
    c.fill = PatternFill('solid', fgColor='F5F5F5')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    # Sutun basliklari
    cols = ['#', 'KLASOR ADI\n(fotograflar buraya)', 'PROJE ADI\n(sitede gorunur)',
            'KONUM\n(sehir/ilce)', 'ALAN (m2)', 'YIL',
            'ETIKET\n(konut/ticari/endustriyel)', 'ACIKLAMA\n(sitede alt yazi)']
    ws.row_dimensions[3].height = 42
    for i, h in enumerate(cols, 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = Font(bold=True, size=9, color='FFFFFF', name='Calibri')
        cell.fill = PatternFill('solid', fgColor='1E3A5F' if i > 1 else 'F4A261')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thick()

    # Ornek satir
    ornek = ['1', 'proje-01', 'Villa Projesi - Tekirdag', 'Suleymanpasa/Tekirdag',
             '180', '2024', kat, '140mm hafif celik karkas, Boardex cephe, cift kat alcipan']
    ws.row_dimensions[4].height = 22
    for i, v in enumerate(ornek, 1):
        cell = ws.cell(row=4, column=i, value=v)
        cell.font = Font(size=9, name='Calibri', color='222222')
        cell.fill = PatternFill('solid', fgColor='EBF5FB')
        cell.alignment = Alignment(horizontal='center' if i in [1,5,6] else 'left', vertical='center')
        cell.border = thin()

    # Bos giris satirlari (5-54)
    for r in range(5, 55):
        ws.row_dimensions[r].height = 20
        bg = 'F8FAFB' if r % 2 == 0 else 'FFFFFF'
        for ci in range(1, 9):
            cell = ws.cell(row=r, column=ci, value='')
            cell.fill = PatternFill('solid', fgColor=bg)
            cell.border = thin()
            cell.alignment = Alignment(horizontal='center' if ci in [1,5,6] else 'left', vertical='center')
        ws.cell(row=r, column=7).value = kat
        ws.cell(row=r, column=7).font = Font(size=9, color='AAAAAA', name='Calibri')

    # Alt bilgi
    ws.merge_cells('A55:H55')
    c = ws['A55']
    c.value = 'Klasor yolu: WEB SITESI FOTOLARI/' + kat + '/[klasor-adi]/ — Proje adi ve aciklama siteye yansiir — GUNCELLE.bat ile guncelle'
    c.font = Font(italic=True, size=8, color='666666')
    c.fill = PatternFill('solid', fgColor='F0F0F0')
    c.alignment = Alignment(horizontal='center', vertical='center')

    ws.freeze_panes = 'B4'
    ws.sheet_view.showGridLines = False

    path = os.path.join(BASE, kat, 'PROJE_BILGI_FOYU_' + kat.upper() + '.xlsx')
    wb.save(path)
    print('Olusturuldu:', path)

print('Tum Excel foleri hazir!')
